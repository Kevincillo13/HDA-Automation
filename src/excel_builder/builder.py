from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.common.config import get_settings
from src.processing.logic import classify_mail_group, normalize_numeric_code

try:
    import win32com.client  # type: ignore[import-not-found]
except ImportError:  # pragma: no cover - depends on Windows environment
    win32com = None  # type: ignore[assignment]


import sys

def get_resource_path(relative_path):
    """Obtiene la ruta absoluta para recursos, compatible con PyInstaller."""
    if hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS) / relative_path
    return Path.cwd() / relative_path

class AP15Builder:
    """Genera AP15 a partir de template.xlsx y exporta CSVs desde ese layout."""

    def __init__(self, output_dir: str, template_path: str | None = None) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        settings = get_settings()
        
        # Intentar resolver la ruta de la plantilla de forma robusta
        raw_path = Path(template_path or settings.template_path)
        if raw_path.exists():
            self.template_path = raw_path
        else:
            # Si no existe en la ruta directa, buscar en el paquete del EXE
            bundled_path = get_resource_path(raw_path)
            self.template_path = bundled_path

    def build(
        self,
        records: list[dict[str, Any]],
        file_suffix: str = "",
    ) -> list[str]:
        grouped_records: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
        for record in records:
            mail_group = classify_mail_group(self._clean(record.get("CompanyCode")))
            vendor_num = self._clean(record.get("VendorNum"))
            currency = self._clean(record.get("Currency"))
            grouped_records[(mail_group, vendor_num, currency)].append(record)

        if not self.template_path.exists():
            raise FileNotFoundError(
                f"AP15 template not found: {self.template_path.resolve()}"
            )

        output_paths: list[str] = []
        suffix = f"_{file_suffix}" if file_suffix else ""
        for (mail_group, vendor_num, currency), grouped in grouped_records.items():
            base_name = f"AP15_{mail_group}_{vendor_num}_{currency}{suffix}"
            workbook_path = self.output_dir / f"{base_name}.xlsx"
            csv_path = self.output_dir / f"{base_name}.csv"

            workbook = load_workbook(self.template_path)
            worksheet = workbook.active
            self._fill_template_rows(worksheet, grouped)
            workbook.save(workbook_path)
            workbook.close()

            self._export_workbook_to_csv(workbook_path, csv_path)
            output_paths.append(str(csv_path))

        return output_paths

    def _fill_template_rows(self, worksheet: Any, records: list[dict[str, Any]]) -> None:
        for index, record in enumerate(records, start=2):
            company_code = self._clean(record.get("CompanyCode"))
            vendor_num = self._clean(record.get("VendorNum"))
            invoice_num = self._clean(record.get("InvoiceNum"))
            invoice_date = self._format_invoice_date(record.get("InvoiceDate"))
            amount = self._format_amount(record.get("Amount", ""))
            currency = self._clean(record.get("Currency"))
            center_value = self._normalize_cost_center(record.get("CostCenter"))
            gl_account = self._normalize_account(record.get("GLAccount"))
            profit_center, cost_center = self._route_center_fields(center_value, gl_account)

            worksheet[f"A{index}"] = company_code
            worksheet[f"B{index}"] = vendor_num
            worksheet[f"C{index}"] = invoice_num
            worksheet[f"D{index}"] = invoice_date
            worksheet[f"E{index}"] = "ITEM"
            worksheet[f"F{index}"] = "DR"
            if amount != "":
                worksheet[f"G{index}"] = float(amount)
                # Use a clean format without thousands separator for SAP compatibility
                worksheet[f"G{index}"].number_format = "0.00"
            else:
                worksheet[f"G{index}"] = ""
            worksheet[f"H{index}"] = currency
            worksheet[f"K{index}"] = company_code
            worksheet[f"L{index}"] = profit_center
            worksheet[f"M{index}"] = cost_center
            worksheet[f"N{index}"] = self._clean(record.get("WBS"))
            worksheet[f"O{index}"] = ""
            worksheet[f"P{index}"] = gl_account
            worksheet[f"Q{index}"] = ""
            worksheet[f"R{index}"] = ""
            worksheet[f"S{index}"] = ""
            worksheet[f"T{index}"] = self._clean(record.get("PayableTo"))
            worksheet[f"U{index}"] = self._clean(record.get("Address"))
            worksheet[f"V{index}"] = ""
            worksheet[f"W{index}"] = self._clean(record.get("City"))
            worksheet[f"X{index}"] = self._clean(record.get("State"))
            worksheet[f"Y{index}"] = self._clean(record.get("Zip"))
            worksheet[f"Z{index}"] = self._clean(record.get("Country"))
            worksheet[f"AA{index}"] = ""
            worksheet[f"AB{index}"] = ""
            worksheet[f"AC{index}"] = ""

    def _export_workbook_to_csv(self, workbook_path: Path, csv_path: Path) -> None:
        if win32com is not None:
            self._export_workbook_to_csv_via_excel(workbook_path, csv_path)
            return

        # Fallback only when Excel COM is unavailable.
        self._export_workbook_to_csv_via_python(workbook_path, csv_path)

    def _export_workbook_to_csv_via_excel(self, workbook_path: Path, csv_path: Path) -> None:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = None
        try:
            workbook = excel.Workbooks.Open(str(workbook_path.resolve()))
            target_path = str(csv_path.resolve())
            if csv_path.exists():
                csv_path.unlink()
            # 62 = xlCSVUTF8. This matches the manual "CSV UTF-8" save flow well.
            workbook.SaveAs(target_path, FileFormat=62, Local=True)
        finally:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
            excel.Quit()

    def _export_workbook_to_csv_via_python(self, workbook_path: Path, csv_path: Path) -> None:
        workbook = load_workbook(workbook_path, data_only=True)
        worksheet = workbook.active
        max_row = self._find_last_non_empty_row(worksheet)
        max_column = 29

        with csv_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.writer(csv_file, lineterminator="\r\n")
            for row_index in range(1, max_row + 1):
                row_values: list[str] = []
                for column_index in range(1, max_column + 1):
                    value = worksheet.cell(row=row_index, column=column_index).value
                    row_values.append(self._serialize_cell_value(value))
                writer.writerow(row_values)

        workbook.close()

    def _find_last_non_empty_row(self, worksheet: Any) -> int:
        last_row = 1
        for row_index in range(1, worksheet.max_row + 1):
            if any(
                worksheet.cell(row=row_index, column=column_index).value not in (None, "")
                for column_index in range(1, 30)
            ):
                last_row = row_index
        return last_row

    def _serialize_cell_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            return f"{value:.2f}"
        return str(value)

    def _normalize_cost_center(self, value: Any) -> str:
        cleaned = self._clean(value)
        if cleaned in {"", "Attached"}:
            return ""
        normalized = normalize_numeric_code(cleaned, width=10)
        return "" if normalized == "Empty" else normalized

    def _normalize_account(self, value: Any) -> str:
        cleaned = self._clean(value).replace(" ", "")
        return "" if cleaned == "Empty" else cleaned

    def _route_center_fields(self, center_value: str, account_value: str) -> tuple[str, str]:
        normalized_account = self._clean(account_value).upper()
        if not center_value:
            return "", ""

        profit_prefixes = ("11", "12", "13", "P1", "P2", "P3")
        cost_prefixes = ("14", "15", "16", "P4", "P5", "P6")

        if normalized_account.startswith(profit_prefixes):
            return center_value, ""
        if normalized_account.startswith(cost_prefixes):
            return "", center_value

        return "", center_value

    def _format_invoice_date(self, value: Any) -> str:
        cleaned = self._clean(value)
        if not cleaned:
            return ""

        parts = cleaned.split("/")
        if len(parts) != 3:
            return cleaned

        try:
            month = str(int(parts[0]))
            day = str(int(parts[1]))
            year = str(int(parts[2]))
            return f"{month}/{day}/{year}"
        except ValueError:
            return cleaned

    def _format_amount(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        try:
            return f"{float(value):.2f}"
        except (TypeError, ValueError):
            return self._clean(value)

    def _clean(self, value: Any) -> str:
        if value is None:
            return ""
        cleaned = str(value).strip()
        return "" if cleaned == "Empty" else cleaned

    def parse_amount(self, value: Any) -> tuple[float, list[str]]:
        """Tries to parse amount safely. Returns (value, errors)."""
        raw_val = self._clean(value).replace("$", "").replace(" ", "")
        if not raw_val:
            return 0.0, ["Amount is missing."]
        
        # Check for multiple decimal-like separators or strange chars
        if raw_val.count(".") > 1 and raw_val.count(",") > 1:
             return 0.0, [f"Ambiguous amount format (multiple separators): '{raw_val}'"]

        normalized = raw_val
        # Case: 1,000.00 (Standard US)
        if "," in normalized and "." in normalized:
            if normalized.rfind(".") > normalized.rfind(","):
                normalized = normalized.replace(",", "")
            else:
                # Case: 1.000,00 (European/Latam)
                # In USD/CAD context, this is risky. Let's handle it but could be rejected if preferred.
                normalized = normalized.replace(".", "").replace(",", ".")
        
        # Case: 4.000 (Could be 4 or 4000)
        elif "." in normalized and normalized.count(".") == 1:
            parts = normalized.split(".")
            if len(parts[-1]) != 2 and len(parts[-1]) != 3:
                # Ambiguous if not .XX or .XXX
                pass 
        
        # Case: 4,000 or 4,50
        elif "," in normalized and normalized.count(",") == 1:
            parts = normalized.split(",")
            if len(parts[-1]) == 3:
                # Likely thousands
                normalized = normalized.replace(",", "")
            elif len(parts[-1]) in [1, 2]:
                # Likely decimal
                normalized = normalized.replace(",", ".")

        try:
            val = float(normalized)
            return val, []
        except ValueError:
            return 0.0, [f"Invalid amount format: '{raw_val}'"]

    def merge_csvs(self, csv_paths: list[str], output_name: str) -> str:
        """Merges multiple AP15 CSV files into a single one, keeping only the first header."""
        if not csv_paths:
            raise ValueError("No CSV paths provided for merging.")

        output_path = self.output_dir / f"{output_name}.csv"
        header_written = False

        with output_path.open("w", encoding="utf-8-sig", newline="") as outfile:
            writer = csv.writer(outfile, lineterminator="\r\n")
            for path_str in csv_paths:
                path = Path(path_str)
                if not path.exists():
                    continue
                
                with path.open("r", encoding="utf-8-sig", newline="") as infile:
                    reader = csv.reader(infile)
                    try:
                        header = next(reader)
                        if not header_written:
                            writer.writerow(header)
                            header_written = True
                        
                        for row in reader:
                            if any(cell.strip() for cell in row):
                                writer.writerow(row)
                    except StopIteration:
                        continue

        return str(output_path)
