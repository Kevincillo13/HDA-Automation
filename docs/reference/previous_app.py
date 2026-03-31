import sys
import os
import re
import importlib
import pdfplumber
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                               QPushButton, QFileDialog, QTextEdit, QLabel, QTabWidget, 
                               QListWidget, QListWidgetItem, QProgressBar, QFrame, QMessageBox)
from PySide6.QtCore import QThread, Signal, Qt
from PySide6.QtGui import QIcon, QDragEnterEvent, QDropEvent
import openpyxl
from openpyxl.styles import PatternFill
import shutil

def resource_path(relative_path):
    """ Obtiene la ruta absoluta al recurso, funciona para dev y para PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def close_pyinstaller_splash():
    """Cierra el splash solo cuando la app corre empaquetada con PyInstaller."""
    if not getattr(sys, "frozen", False):
        return

    try:
        splash_module = importlib.import_module("pyi_splash")
    except ImportError:
        return

    splash_module.close()

class VisualExtractor(QThread):
    log_signal = Signal(str)
    progress_signal = Signal(int)
    finished_signal = Signal()

    def __init__(self, pdf_files):
        super().__init__()
        self.pdf_files = pdf_files
        self.US_STATES = {
            'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR', 'california': 'CA',
            'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE', 'florida': 'FL', 'georgia': 'GA',
            'hawaii': 'HI', 'idaho': 'ID', 'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA',
            'kansas': 'KS', 'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
            'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS', 'missouri': 'MO',
            'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV', 'new hampshire': 'NH', 'new jersey': 'NJ',
            'new mexico': 'NM', 'new york': 'NY', 'north carolina': 'NC', 'north dakota': 'ND', 'ohio': 'OH',
            'oklahoma': 'OK', 'oregon': 'OR', 'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
            'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT', 'vermont': 'VT',
            'virginia': 'VA', 'washington': 'WA', 'west virginia': 'WV', 'wisconsin': 'WI', 'wyoming': 'WY'
        }
        self.CAN_PROVINCES = {
            'alberta': 'AB', 'british columbia': 'BC', 'manitoba': 'MB', 'new brunswick': 'NB',
            'newfoundland': 'NL', 'nova scotia': 'NS', 'ontario': 'ON', 'prince edward island': 'PE',
            'quebec': 'QC', 'saskatchewan': 'SK'
        }

    def parse_city_state_zip(self, text):
        """Separa 'Mason OH 45040' o 'Whitby, ON L1R 2S7, Canada' en City, State, Zip, Country"""
        if not text or text == "Empty":
            return "Empty", "Empty", "Empty", "US" # Default
            
        clean_text = text.replace(',', ' ').strip()
        # Eliminar explícitamente "Canada" o "United States" si vienen al final
        clean_text = re.sub(r'\b(Canada|USA|United States)\b', '', clean_text, flags=re.IGNORECASE).strip()
        
        parts = clean_text.split()
        
        zip_code = "Empty"
        state = "Empty"
        city = "Empty"
        country = "US" # Default
        
        if not parts: return city, state, zip_code, country

        # 1. Detectar ZIP
        # Caso Canadá Pegado: H9R1A3
        if len(parts) >= 1 and re.match(r'^[A-Za-z]\d[A-Za-z]\d[A-Za-z]\d$', parts[-1]):
            zip_code = f"{parts[-1][:3]} {parts[-1][3:]}".upper()
            country = "CA" # <-- FIX APLICADO: Si el formato es canadiense, fuerza el país
            parts = parts[:-1]
        # Caso Canadá Separado: M2J 0B3
        elif len(parts) >= 2 and re.match(r'^[A-Za-z]\d[A-Za-z]$', parts[-2]) and re.match(r'^\d[A-Za-z]\d$', parts[-1]):
            zip_code = f"{parts[-2]} {parts[-1]}".upper()
            country = "CA" # <-- FIX APLICADO: Si el formato es canadiense, fuerza el país
            parts = parts[:-2]
        # Caso USA: 45040 o 45040-1234
        elif len(parts) >=1 and re.match(r'^\d{5}(-\d{4})?$', parts[-1]):
            zip_code = parts[-1]
            parts = parts[:-1]
        
        if not parts: return city, state, zip_code, country

        # 2. Detectar Estado/Provincia
        potential_state = parts[-1].lower()
        
        if len(parts[-1]) == 2 and parts[-1].isalpha():
            state_code = parts[-1].upper()
            state = state_code
            parts = parts[:-1]
            if state_code in self.US_STATES.values(): country = "US"
            elif state_code in self.CAN_PROVINCES.values(): country = "CA"
            
        elif potential_state in self.US_STATES:
            state = self.US_STATES[potential_state]
            country = "US"
            parts = parts[:-1]
        elif potential_state in self.CAN_PROVINCES:
            state = self.CAN_PROVINCES[potential_state]
            country = "CA"
            parts = parts[:-1]
            
        # 3. Lo que sobra es la Ciudad
        city = " ".join(parts).title()
        return city, state, zip_code, country

    def get_coords(self, page, text_fragment, after_y=0):
        try:
            matches = page.search(re.escape(text_fragment), regex=True, case=False)
            for m in matches:
                if m['top'] >= after_y:
                    return m['x1'], m['top'], m['bottom'], m['x0']
        except Exception:
            pass
        return None

    def extract_visual_field(self, page, label_start, label_end_list, y_tolerance=5, right_margin=0, left_margin=0, after_y=0):
        start_coords = self.get_coords(page, label_start, after_y)
        if not start_coords: return "Empty"

        start_x, start_top, start_bottom, _ = start_coords
        start_x += left_margin + 2 
        
        limit_x = page.width
        original_limit = page.width
        
        for label_end in label_end_list:
            end_coords = self.get_coords(page, label_end, after_y)
            if end_coords:
                _, end_top, _, end_left_edge = end_coords
                if abs(end_top - start_top) < 10 and end_left_edge > start_x:
                    if end_left_edge < limit_x:
                        limit_x = end_left_edge

        if limit_x < original_limit:
            limit_x -= right_margin

        try:
            crop = page.crop((start_x, start_top - 2, limit_x, start_bottom + 2))
            text = crop.extract_text()
            if text:
                text = text.replace(label_start, "").strip()
                text = text.lstrip(" :")
                return text if text else "Empty"
        except:
            return "Empty"
        return "Empty"

    def run(self):
        grouped_data = {} 
        total_files = len(self.pdf_files)

        for i, pdf_path in enumerate(self.pdf_files):
            try:
                filename = os.path.basename(pdf_path)
                with pdfplumber.open(pdf_path) as pdf:
                    page = pdf.pages[0]
                    width = page.width
                    height = page.height

                    # 1. ANCLAS VISUALES CLAVE
                    header_bottom = 200
                    coords_subj = self.get_coords(page, "Subject:")
                    if coords_subj: header_bottom = coords_subj[1]

                    coords_comp = self.get_coords(page, "Company:")
                    company_top = coords_comp[1] if coords_comp else 400

                    # 2. EXTRACCIÓN DE ENCABEZADOS (Header)
                    raw_contact = self.extract_visual_field(page, "Contact:", ["Site:", "Created:", "Priority:"])
                    id_match = re.search(r'(\d{6,7}G)', raw_contact)
                    clean_id = id_match.group(1) if id_match else "Empty"
                    clean_contact = raw_contact.replace(clean_id, "").strip()
                    if clean_id == "Empty":
                        header_dump = page.crop((0,0,width, header_bottom)).extract_text()
                        id_match_glob = re.search(r'(\d{6,7}G)', header_dump)
                        if id_match_glob: clean_id = id_match_glob.group(1)

                    data = {
                        "Id": clean_id,
                        "Contact": clean_contact,
                        "Site": self.extract_visual_field(page, "Site:", ["Created:", "Priority:", "Assigned"]),
                        "Created": self.extract_visual_field(page, "Created:", ["Priority:", "Assigned"]),
                        "Priority": self.extract_visual_field(page, "Priority:", ["Assigned"], right_margin=5),
                        "Assigned to": self.extract_visual_field(page, "Assigned to:", ["Subject:"]),
                        "Subject": self.extract_visual_field(page, "Subject:", ["Problem"]),
                    }
                    
                    if data["Created"].endswith("P"): data["Created"] = data["Created"][:-1].strip()

                    # 3. EXTRACCIÓN PROBLEM / SOLUTION
                    y_start_prob = header_bottom + 20
                    y_end_prob = company_top - 5

                    split_x = width * 0.5 
                    coords_sol = self.get_coords(page, "Solution")
                    if not coords_sol: coords_sol = self.get_coords(page, "Solution:")
                    if coords_sol: split_x = coords_sol[3] - 10

                    try:
                        prob_crop = page.crop((0, y_start_prob, split_x, y_end_prob))
                        sol_crop = page.crop((split_x, y_start_prob, width, y_end_prob))
                        
                        p_text = prob_crop.extract_text() or ""
                        s_text = sol_crop.extract_text() or ""
                        
                        data["Problem"] = p_text.replace("Problem", "").replace("Solu", "").strip()
                        data["Solution"] = s_text.replace("Solution", "").replace("ution", "").strip()

                        if not data["Problem"]: data["Problem"] = "Empty"
                        if not data["Solution"]: data["Solution"] = "Empty"
                    except:
                        data["Problem"] = "Empty"
                        data["Solution"] = "Empty"

                    # 4. EXTRACCIÓN DE DATOS FINANCIEROS (Middle)
                    mid_labels = ["Vendor", "Invoice Date", "Invoice Number", "POR", "Amount", "Currency", "Payable", "Address", "City"]
                    
                    data["Company"] = self.extract_visual_field(page, "Company:", mid_labels, after_y=header_bottom)
                    data["Vendor #"] = self.extract_visual_field(page, "Vendor #:", mid_labels, right_margin=5, after_y=company_top)
                    data["Invoice Date"] = self.extract_visual_field(page, "Invoice Date:", mid_labels, right_margin=5, after_y=company_top)
                    data["Invoice Number"] = self.extract_visual_field(page, "Invoice Number:", mid_labels, after_y=company_top)
                    data["POR #"] = self.extract_visual_field(page, "POR #:", mid_labels, right_margin=5, after_y=company_top)
                    data["Amount"] = self.extract_visual_field(page, "Amount:", ["Currency", "Payable"], right_margin=5, after_y=company_top)
                    data["Currency"] = self.extract_visual_field(page, "Currency:", ["Payable", "Address", "City"], after_y=company_top)
                    
                    coords_payable = self.get_coords(page, "Payable To:", after_y=company_top)
                    coords_address = self.get_coords(page, "Address:", after_y=company_top)

                    if coords_payable:
                        if coords_address and abs(coords_payable[1] - coords_address[1]) < 10:
                            data["Payable to"] = self.extract_visual_field(page, "Payable To:", ["Address:"], right_margin=5, after_y=company_top)
                        else:
                            end_y = coords_address[1] if coords_address else coords_payable[2] + 30
                            pay_crop = page.crop((coords_payable[0] + 2, coords_payable[1] - 2, width, end_y - 2))
                            text = pay_crop.extract_text()
                            if text: data["Payable to"] = text.replace("Payable To:", "").strip()
                            else: data["Payable to"] = "Empty"
                    else:
                        data["Payable to"] = "Empty"

                    if not data.get("Payable to") or data.get("Payable to") == "": data["Payable to"] = "Empty"

                    coords_city = self.get_coords(page, "City/State")
                    right_limit = (coords_city[3] - 5) if coords_city else width

                    if coords_address:
                        limit_y = coords_city[1] if coords_city else (coords_address[1] + 30)
                        addr_crop = page.crop((coords_address[0] + 2, coords_address[1], right_limit, limit_y + 10))
                        data["Address"] = addr_crop.extract_text().replace("Address:", "").lstrip(" :").strip()
                        
                        if coords_city:
                             city_crop = page.crop((coords_city[0] + 5, coords_city[1], width, limit_y + 20))
                             data["City/State"] = city_crop.extract_text().replace("City/State", "").replace("Zip:", "").strip()
                        else:
                             data["City/State"] = "Empty"

                    # CORRECCIÓN DE SALTO DE LÍNEA EXTREMO EN DIRECCIONES
                    address_val = data.get("Address", "Empty")
                    city_state_val = data.get("City/State", "Empty")
                    
                    if address_val == "Empty" and " " in city_state_val:
                        parts = city_state_val.split()
                        if parts[0].isdigit() or parts[0].lower() in ["p.o.", "po", "box"]:
                            data["Address"] = city_state_val
                            data["City/State"] = "Empty" 

                    # CORRECCIÓN ESPECÍFICA EYEMED (NOMBRE ROTO)
                    if "Employees Health Trust" in data["City/State"]:
                         data["Payable to"] += " Employees Health Trust"
                         data["City/State"] = data["City/State"].replace("Employees Health Trust", "").strip()
                    
                    data["Vendor Contact"] = self.extract_visual_field(page, "Vendor Contact", ["Cost/Profit"])

                    # 5. TABLA CONTABLE ROBUSTA
                    data.update({"Cost/Profit center": "Empty", "GL Account": "Empty", "WBS Element": "Empty", "Distribution AMT": "Empty", "Brand code": "Empty"})
                    
                    tables = page.extract_tables()
                    target_table = None
                    for t in tables:
                        t_str = str(t).lower()
                        if "gl account" in t_str or "cost/profit" in t_str:
                            target_table = t
                            break
                    
                    if target_table:
                        for row in target_table:
                            clean_row = [str(c).replace('\n',' ').strip() for c in row if c is not None]
                            row_str = "".join(clean_row).lower()
                            
                            if any(char.isdigit() for char in row_str) and "gl account" not in row_str:
                                if len(clean_row) >= 1: data["Cost/Profit center"] = clean_row[0]
                                if len(clean_row) >= 2: data["GL Account"] = clean_row[1]
                                
                                remaining = clean_row[2:]
                                for item in remaining:
                                    item_clean = item.lower().replace("usd","").replace("cad","").strip()
                                    
                                    if len(item_clean) >= 9 and item_clean.isdigit():
                                        data["WBS Element"] = item_clean
                                    elif "." in item_clean or "," in item_clean or (item_clean.isdigit() and len(item_clean) < 8):
                                        data["Distribution AMT"] = item_clean
                                    elif len(item_clean) > 0:
                                        data["Brand code"] = item_clean
                    
                    text_full = page.extract_text()
                    pay_match = re.search(r'Payment Method:\s*(.*?)(?=\n|"Check")', text_full)
                    if pay_match: data["Payment method"] = pay_match.group(1).strip()
                    else: data["Payment method"] = "OneTime Check"

                    # --- APLICAR REGLAS DE NEGOCIO PARA EXCEL ---
                    comp_text = data.get("Company", "")
                    comp_code_match = re.match(r'^([A-Za-z0-9]+)', comp_text)
                    comp_code = comp_code_match.group(1) if comp_code_match else "Empty"
                    
                    vendor_num = "8000001"
                    if "E100" in comp_code.upper() or "E1OO" in comp_code.upper():
                        vendor_num = "900000"
                    elif comp_code in ["1000", "2000"]:
                        vendor_num = "900010"
                    
                    inv_num = data.get("Invoice Number")
                    if not inv_num or inv_num == "Empty": inv_num = data.get("Id")
                    
                    inv_date = data.get("Invoice Date")
                    if not inv_date or inv_date == "Empty": inv_date = data.get("Created")
                    if inv_date and inv_date != "Empty":
                        inv_date = inv_date.split(" ")[0]

                    amt_str = str(data.get("Amount", "0")).replace(",", "").replace("$", "").strip()
                    try:
                        amount_val = float(amt_str)
                    except:
                        amount_val = 0.0

                    city, state, zip_c, country_from_address = self.parse_city_state_zip(data.get("City/State", ""))
                    
                    curr = data.get("Currency", "USD").upper()
                    if country_from_address != "US":
                        country = country_from_address
                    else:
                        country = "CA" if "CAD" in curr else "US"

                    processed_row = {
                        "CompanyCode": comp_code,
                        "VendorNum": vendor_num,
                        "InvoiceNum": inv_num,
                        "InvoiceDate": inv_date,
                        "Amount": amount_val,
                        "Currency": curr,
                        "CostCenter": data.get("Cost/Profit center", "Empty"),
                        "GLAccount": data.get("GL Account", "Empty"),
                        "PayableTo": data.get("Payable to", "Empty").replace('\n', ' ').strip(),
                        "Address": data.get("Address", "Empty").replace('\n', ' ').strip(),
                        "City": city,
                        "State": state,
                        "Zip": zip_c,
                        "Country": country
                    }

                    group_key = (vendor_num, curr)
                    if group_key not in grouped_data:
                        grouped_data[group_key] = []
                    grouped_data[group_key].append(processed_row)

                    self.log_signal.emit(f"--- REPORTE: {filename} ---")
                    for k, v in data.items():
                        clean_v = str(v).replace('\n', ' ').replace("  ", " ").strip()
                        self.log_signal.emit(f"{k}: {clean_v}")
                    self.log_signal.emit("\n")
                
                self.progress_signal.emit(int((i + 1) / total_files * 100))

            except Exception as e:
                self.log_signal.emit(f"ERROR CRÍTICO EN {pdf_path}: {e}")
                import traceback
                traceback.print_exc()

        # --- GENERAR EXCELS POR VENDOR ---
        if grouped_data:
            base_dir = os.path.dirname(self.pdf_files[0])
            template_path = resource_path(os.path.join("templates", "template.xlsx"))
            
            if not os.path.exists(template_path):
                self.log_signal.emit(f"\n[ERROR] No se encontró el template en: {template_path}")
                self.finished_signal.emit()
                return

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for (v_num, curr), rows in grouped_data.items():
                try:
                    output_filename = f"AP15_{v_num}_{curr}.xlsx"
                    output_path = os.path.join(base_dir, output_filename)
                    
                    shutil.copy(template_path, output_path)
                    
                    wb = openpyxl.load_workbook(output_path)
                    ws = wb.active
                    start_row = 2
                    
                    for i, row_data in enumerate(rows):
                        r = start_row + i
                        
                        ws[f"A{r}"] = row_data["CompanyCode"]
                        ws[f"B{r}"] = row_data["VendorNum"]
                        ws[f"C{r}"] = row_data["InvoiceNum"]
                        ws[f"D{r}"] = row_data["InvoiceDate"]
                        ws[f"E{r}"] = "ITEM"
                        ws[f"F{r}"] = "DR"
                        ws[f"G{r}"] = row_data["Amount"]
                        ws[f"G{r}"].number_format = '#,##0.00'
                        ws[f"H{r}"] = row_data["Currency"]
                        ws[f"K{r}"] = row_data["CompanyCode"]
                        
                        # Cost Center (Limpieza y Padding 0s)
                        raw_cc = str(row_data["CostCenter"]).replace(" ", "").strip()
                        if raw_cc.lower() == "attached" or raw_cc.lower() == "empty":
                            cc = "Empty"
                        elif raw_cc.isdigit() or re.match(r'^[0-9\-]+$', raw_cc):
                            if "-" not in raw_cc and raw_cc.isdigit():
                                cc = raw_cc.zfill(10)
                            else:
                                cc = raw_cc
                        else:
                            cc = raw_cc
                            
                        ws[f"M{r}"] = cc
                        if cc == "Empty": ws[f"M{r}"].fill = yellow_fill

                        # GL Account (Quitar espacios internos y validación)
                        raw_gl = str(row_data["GLAccount"]).replace(" ", "").strip()
                        if raw_gl.lower() == "attached" or raw_gl.lower() == "empty":
                            gl = "Empty"
                        else:
                            gl = raw_gl
                            
                        ws[f"P{r}"] = gl
                        if gl == "Empty" or (gl.isdigit() and len(gl) != 10): 
                            ws[f"P{r}"].fill = yellow_fill

                        if gl.startswith("P") and str(row_data["VendorNum"]) != "8000001":
                            ws[f"A{r}"].fill = yellow_fill
                            ws[f"P{r}"].fill = yellow_fill

                        ws[f"T{r}"] = row_data["PayableTo"]
                        ws[f"U{r}"] = row_data["Address"]
                        ws[f"W{r}"] = row_data["City"]
                        ws[f"X{r}"] = row_data["State"]
                        ws[f"Y{r}"] = row_data["Zip"]
                        if row_data["Zip"] == "Empty": ws[f"Y{r}"].fill = yellow_fill
                        ws[f"Z{r}"] = row_data["Country"]
                        
                        if row_data["Amount"] == 0.0: ws[f"G{r}"].fill = yellow_fill

                    wb.save(output_path)
                    self.log_signal.emit(f"\n[EXITO] Generado: {output_filename}")

                except Exception as e:
                    self.log_signal.emit(f"\n[ERROR] Falló al generar Excel para Vendor {v_num} ({curr}): {e}")

        self.finished_signal.emit()

class DragDropArea(QFrame):
    files_dropped = Signal(list)
    clicked = Signal()

    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self.setMinimumHeight(200)
        self.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Sunken)
        self.setStyleSheet("""
            QFrame {
                border: 2px dashed #aaa;
                border-radius: 10px;
                background-color: #f9f9f9;
            }
            QFrame:hover {
                border-color: #1D4F91;
                background-color: #eef4ff;
            }
        """)
        
        layout = QVBoxLayout(self)
        self.label = QLabel("Arrastra tus archivos PDF aquí\no haz clic para seleccionar")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label.setStyleSheet("color: #666; font-size: 16px; border: none; background: transparent;")
        layout.addWidget(self.label)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        files = []
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith(".pdf"):
                files.append(file_path)
        if files:
            self.files_dropped.emit(files)
            
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()

class AP15App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("HDA Automation Tool (BETA v0.2) - PDF to AP15")
        self.resize(900, 650)
        
        icon_path = resource_path("icon.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        
        self.setStyleSheet("""
            QMainWindow { background-color: #f4f6f9; }
            QTabWidget::pane { border: 1px solid #e0e0e0; background: white; border-radius: 5px; }
            QTabBar::tab { background: #e0e0e0; padding: 10px 20px; border-top-left-radius: 5px; border-top-right-radius: 5px; margin-right: 2px; font-weight: bold; color: #555; }
            QTabBar::tab:selected { background: #1D4F91; color: white; }
            QPushButton { background-color: #1D4F91; color: white; border-radius: 5px; padding: 10px; font-weight: bold; font-size: 14px; }
            QPushButton:hover { background-color: #2c5d94; }
            QPushButton:disabled { background-color: #ccc; }
            QListWidget { border: 1px solid #dcdcdc; border-radius: 5px; padding: 5px; background-color: white; font-size: 14px; }
            QProgressBar { border: 1px solid #dcdcdc; border-radius: 5px; text-align: center; background-color: white; height: 20px; }
            QProgressBar::chunk { background-color: #6CACE4; border-radius: 5px; }
            QLabel { color: #333; }
        """)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        header_layout = QHBoxLayout()
        title = QLabel("HDA Automation Tool")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #1D4F91;")
        header_layout.addWidget(title)
        header_layout.addStretch()
        main_layout.addLayout(header_layout)

        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        tab1 = QWidget()
        tab1_layout = QVBoxLayout(tab1)
        
        self.drop_area = DragDropArea()
        self.drop_area.files_dropped.connect(self.add_files)
        self.drop_area.clicked.connect(self.select_files)
        tab1_layout.addWidget(self.drop_area)

        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QListWidget.SelectionMode.NoSelection)
        self.file_list.setSpacing(4)
        tab1_layout.addWidget(QLabel("Archivos seleccionados:"))
        tab1_layout.addWidget(self.file_list)

        btn_layout = QHBoxLayout()
        self.btn_process = QPushButton("Generar Reportes AP15")
        self.btn_process.clicked.connect(self.run_process)
        btn_layout.addWidget(self.btn_process)
        tab1_layout.addLayout(btn_layout)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        tab1_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("Listo para procesar.")
        self.status_label.setStyleSheet("font-style: italic; color: #666;")
        tab1_layout.addWidget(self.status_label)

        self.tabs.addTab(tab1, "PDF to AP15")

        tab2 = QWidget()
        tab2_layout = QVBoxLayout(tab2)
        lbl_coming_soon = QLabel("Funcionalidad End 2 End\nPróximamente")
        lbl_coming_soon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_coming_soon.setStyleSheet("font-size: 20px; color: #aaa; font-weight: bold;")
        tab2_layout.addWidget(lbl_coming_soon)
        self.tabs.addTab(tab2, "End 2 End")

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Seleccionar PDFs", "", "PDF Files (*.pdf)")
        if files:
            self.add_files(files)

    def add_files(self, files):
        for f in files:
            items = [self.file_list.item(i).data(Qt.UserRole) for i in range(self.file_list.count())]
            if f not in items:
                item = QListWidgetItem(self.file_list)
                item.setData(Qt.UserRole, f)
                
                widget = QWidget()
                h_layout = QHBoxLayout(widget)
                h_layout.setContentsMargins(15, 12, 15, 12)
                
                lbl_name = QLabel(os.path.basename(f))
                lbl_name.setStyleSheet("font-weight: bold;")
                
                btn_remove = QPushButton("✕")
                btn_remove.setFixedSize(30, 30)
                btn_remove.setCursor(Qt.CursorShape.PointingHandCursor)
                btn_remove.setStyleSheet("""
                    QPushButton { background-color: transparent; color: #aaa; border: none; font-size: 16px; }
                    QPushButton:hover { color: #e74c3c; background-color: #fadbd8; border-radius: 15px; }
                """)
                btn_remove.clicked.connect(lambda _, it=item: self.remove_file(it))
                
                h_layout.addWidget(lbl_name)
                h_layout.addStretch()
                h_layout.addWidget(btn_remove)
                
                item.setSizeHint(widget.sizeHint())
                self.file_list.setItemWidget(item, widget)
                
        self.status_label.setText(f"{self.file_list.count()} archivos cargados.")

    def remove_file(self, item):
        row = self.file_list.row(item)
        self.file_list.takeItem(row)
        self.status_label.setText(f"{self.file_list.count()} archivos cargados.")

    def run_process(self):
        count = self.file_list.count()
        if count == 0:
            QMessageBox.warning(self, "Sin archivos", "Por favor agrega archivos PDF primero.")
            return

        files = [self.file_list.item(i).data(Qt.UserRole) for i in range(count)]
        
        self.processing_errors = []
        self.progress_bar.setValue(0)
        self.btn_process.setEnabled(False)
        self.drop_area.setEnabled(False)
        self.file_list.setEnabled(False)
        self.status_label.setText("Procesando...")

        self.worker = VisualExtractor(files)
        self.worker.log_signal.connect(self.handle_log)
        self.worker.progress_signal.connect(self.progress_bar.setValue)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def handle_log(self, msg):
        if "[ERROR]" in msg or "ERROR CRÍTICO" in msg:
            self.processing_errors.append(msg)
        if "[EXITO]" in msg:
            self.status_label.setText(msg.strip())
            
        try:
            with open("beta_debug_log.txt", "a", encoding="utf-8") as f:
                f.write(msg + "\n")
        except:
            pass

    def on_finished(self):
        self.btn_process.setEnabled(True)
        self.drop_area.setEnabled(True)
        self.file_list.setEnabled(True)
        self.progress_bar.setValue(100)
        self.status_label.setText("Proceso completado. Revisa la carpeta de origen.")
        
        if self.processing_errors:
            error_text = "\n".join(self.processing_errors)
            QMessageBox.warning(self, "Proceso con Alertas", f"El proceso terminó, pero hubo errores:\n\n{error_text}")
        else:
            QMessageBox.information(self, "Éxito", "¡Todo listo!\nSe han generado los archivos Excel correctamente.")

if __name__ == "__main__":
    try:
        import ctypes
        myappid = 'essilorluxottica.hda.automation.tool.v0.2'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    except:
        pass

    app = QApplication(sys.argv)
    
    window = AP15App()
    window.show()
    
    close_pyinstaller_splash()
    
    sys.exit(app.exec())
